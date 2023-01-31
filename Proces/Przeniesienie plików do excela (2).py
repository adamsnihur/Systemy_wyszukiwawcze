# Poprawiony kod:
import csv
import openpyxl

# Otwieramy plik csv i wczytujemy dane do zmiennej csv
with open('/HTML/templates/dane_bez_spacji.csv', 'r') as csv_file:
    csv_reader = csv.reader(csv_file, delimiter='/')

# Otwieramy plik Excel.xlsx i tworzymy arkusz
    wb = openpyxl.Workbook()
    sheet = wb.active

# W wierszu pierwszym wypisujemy nazwy kolumn
    sheet.cell(row=1, column=1).value = "data"
    sheet.cell(row=1, column=2).value = "sklep"
    sheet.cell(row=1, column=3).value = "cena"
    sheet.cell(row=1, column=4).value = "produkt"

# Wczytujemy dane z pliku csv do pliku Excel.xlsx
    for row, data in enumerate(csv_reader):
        for col, value in enumerate(data):
            sheet.cell(row=row+2, column=col+1).value = value

# Zapisujemy plik Excel.xlsx
    wb.save('/Users/joachimuetake/Lewoniewski projekt/Excel.xlsx')